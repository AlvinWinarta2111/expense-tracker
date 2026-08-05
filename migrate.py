"""
One-off migration: reads Living_Cost_Cikarang.xlsx and inserts every
transaction into Supabase as a row in `entries`.

Usage:
    python migrate.py                # dry run, just prints what it would insert
    python migrate.py --commit       # actually inserts into Supabase

Place a copy of Living_Cost_Cikarang.xlsx in this same folder first,
or pass a path: python migrate.py --file /path/to/file.xlsx --commit
"""

import argparse

import openpyxl
from supabase import create_client
import toml

parser = argparse.ArgumentParser()
parser.add_argument("--file", default="Living_Cost_Cikarang.xlsx")
parser.add_argument("--commit", action="store_true", help="actually write to Supabase")
args = parser.parse_args()

secrets = toml.load(".streamlit/secrets.toml")
sb = create_client(secrets["SUPABASE_URL"], secrets["SUPABASE_KEY"])

cat_rows = sb.table("categories").select("id,name").execute().data
cat_map = {c["name"].strip().lower(): c["id"] for c in cat_rows}

wb = openpyxl.load_workbook(args.file, data_only=True)

to_insert = []
for sheet_index, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    last_date = None
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        item = row[1].value  # column B
        if item is None:
            continue
        if row[2].value is not None:  # column C, date
            last_date = row[2].value

        if item == "Starting Balance" and sheet_index != 0:
            continue  # every month after the first re-injects last month's
            # balance as a fake income row - skip it, the running total
            # already carries forward automatically in the new schema
        category_name = row[3].value  # column D
        credit = row[4].value or 0  # column E
        debit = row[5].value or 0  # column F
        amount = float(credit) - float(debit)

        cat_id = cat_map.get(str(category_name).strip().lower()) if category_name else None
        if cat_id is None:
            print(f"[skip] unknown category '{category_name}' on '{item}' ({sheet_name})")
            continue

        to_insert.append(
            {
                "item": str(item),
                "entry_date": last_date.date().isoformat() if last_date else None,
                "category_id": cat_id,
                "amount": amount,
                "note": f"migrated from {sheet_name}",
            }
        )

print(f"Found {len(to_insert)} entries to migrate.")
for e in to_insert[:5]:
    print(" ", e)
print("  ...")

if args.commit:
    CHUNK = 200
    for i in range(0, len(to_insert), CHUNK):
        sb.table("entries").insert(to_insert[i : i + CHUNK]).execute()
    print(f"Inserted {len(to_insert)} entries into Supabase.")
else:
    print("\nDry run only - nothing written. Re-run with --commit to actually insert.")
